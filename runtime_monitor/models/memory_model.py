
import json
from collections import deque
from functools import partial

#from adios2_tau_reader import adios2_tau_reader
from mpi4py import MPI
import sys
#from pandas.tseries.frequencies import to_offset
#from pandas import Timestamp
import datetime as dt
import os
import numpy
import pandas as pd
from collections import Counter
from collections import OrderedDict
from enum import Enum
import math
import openpyxl
import xlsxwriter
from math import sqrt
import json
from runtime_monitor import helper 
from runtime_monitor import abstract_model

likwid_counters = {}
papi_counters = {
           "Intel(R) Xeon(R) CPU E5-2680 v2 @ 2.80GHz" : {
              "llc_misses": [ "LAST_LEVEL_CACHE_MISSES", "D"],
              "llc_refs": [ "LAST_LEVEL_CACHE_REFERENCES", "D"],
              "ld_stalls": [ "CYCLE_ACTIVITY:STALLS_LDM_PENDING", "D"],
              "inst_ret": ["INSTRUCTIONS_RETIRED", "D"],
              "st_stalls" : ["RESOURCE_STALLS:SB", "D" ],
              "cpu_cyc" : ["ix86arch::UNHALTED_CORE_CYCLES", "D"]
           },
           "POWER9, altivec supported" : {
              "llc_refs": ["perf::LLC-LOADS", "D" ],
              "llc_misses": ["perf::LLC-LOAD-MISSES", "D" ],
              "ld_stalls": ["PM_LD_L3MISS_PEND_CYC", "D"], #PM_CMPLU_STALL_DMISS_L3MISS", "D"],
              "inst_ret": ["PM_INST_CMPL", "D" ],
              "cpu_cyc": ["PM_RUN_CYC", "D"], #perf::PERF_COUNT_SW_CPU_CYCLES", "D" ],
              "gtbw": ["cuda:::metric_nvlink_transmit_throughput:device=0", "D" ],
              "grbw": ["cuda:::metric_nvlink_receive_throughput:device=0", "D" ]
            },
            "Intel(R) Xeon(R) Gold 6248 CPU @ 2.50GHz" : {
               "llc_refs": ["ix86arch::LLC_REFERENCES", "D" ],
               "llc_misses": ["ix86arch::LLC_MISSES", "D"],
               "ld_stalls" : ["CYCLE_ACTIVITY:STALLS_MEM_ANY", "D"],
               "inst_ret": ["ix86arch::INSTRUCTION_RETIRED", "D"],
               "cpu_cyc": ["ix86arch::UNHALTED_CORE_CYCLES", "D"],
               "gtbw": ["GPU: Device DVN PCIe TX Throughput (MB/s)", "D"],
               "grbw": ["GPU: Device DVN PCIe RX Throughput (MB/s)", "D"]
               #"grbw": ["cuda:::metric_nvlink_receive_throughput:device=0", "D" ]
#               "gtbw" : ["GPU: Device 0 NvLink Throughput Raw TX","D"],
#               "grbw": ["GPU: Device 0 NvLink Throughput Raw RX" , "D"]
            }   
       }

metrics = { "Intel(R) Xeon(R) CPU E5-2680 v2 @ 2.80GHz" : [
              "llc_miss_per",
              "ld_stalls_per",
              "ipc"
            ],
            "POWER9, altivec supported" : [
              "llc_miss_per",
              "ld_stalls_per",
              "ipc",
              "gtbw",
              "grbw" 
            ],
            "Intel(R) Xeon(R) Gold 6248 CPU @ 2.50GHz" : [
               "llc_miss_per",
               "ld_stalls_per"
               "ipc"
            ]    
       }

memory_size = {
            "Intel(R) Xeon(R) CPU E5-2680 v2 @ 2.80GHz" : "128",
            "POWER9, altivec supported" : "512",
            "Intel(R) Xeon(R) Gold 6248 CPU @ 2.50GHz": 500, 
        }


class MetricID(Enum):
    RSS = 0
    VMS = 1
    LLCM = 2 
    LLCR = 3
    LDS = 4
    INS = 5
    CYC = 6
    GTBW = 7
    GRBW = 8 

class OrderedCounter(Counter, OrderedDict):
     'Counter that remembers the order elements are first encountered'

     def __repr__(self):
         return '%s(%r)' % (self.__class__.__name__, OrderedDict(self))

     def __reduce__(self):
         return self.__class__, (OrderedDict(self),)


def _fix_ranges_(func, val_l, val_tot):
    if len(val_l) != len(val_tot):
        #print("[Rank ", self.my_rank, "] :",func, " : Length of meaurements donot match!\n", len(val_l), "!=", len(val_tot))
        nlen = min(len(val_l), len(val_tot)) 
        val_l = val_l[:nlen]
        val_tot = val_tot[:nlen]
    return val_l, val_tot
    
def _percentage_(func, nprocs, val_l, val_tot):
    percentage = {}
    max_p = 0
    for i in range(nprocs):
        val_l[i], val_tot[i] = _fix_ranges_(func, val_l[i], val_tot[i])
        percentage[i] = [] 
        percentage[i].extend([ (x/y*100) if y else 0 for x,y in zip(val_l[i] , val_tot[i])])
        if len(percentage[i]) != 0 :
            if max(percentage[i]) == 0:
                print("[Rank ", self.my_rank, "] :","Max percenatge is 0 for index " , i, " in ", func)
        if len(percentage[i]) != 0 and max_p < max(percentage[i]):
             max_p = max(percentage[i])
    return percentage, max_p

def _divide_(func, nprocs, val_l, val_tot):
    div = {}
    max_v = 0
    for i in range(nprocs):
        val_l[i], val_tot[i] = _fix_ranges_(func, val_l[i], val_tot[i])
        div[i] = []
        div[i].extend([ x/y if y else 0 for x,y in zip(val_l[i] , val_tot[i])])
        if len(div[i]) != 0 :
            if max(div[i]) == 0:
                print("[Rank ", self.my_rank, "] :","Max division is 0 for index " , i, " in ", func)
        if len(div[i]) != 0 and max_v < max(div[i]):
             max_v = max(div[i])
    return div, max_v

def _sub_(func, nprocs, val_l):
    sub = {}
    max_v = 0
    for i in range(nprocs):
        sub[i] = []
        sub[i].extend([j-k for k, j in zip(vals[:-1], vals[1:])])
    return sub

class memory(abstract_model.model):
#class memory():

    def __init__(self, config): #adios2_reader_objects, frequency, cpu_model, gpu_n):
        self.iter = 0
        self.hd_counters = {} 
        self.rss_m = ["Memory Footprint (VmRSS) (KB)", "C"]
        self.vms_m = ["Peak Memory Usage Resident Set Size (VmHWM) (KB)", "C"]
        self.metric_func = []
        self.adios2_active_conns = []
        self.r_map = None
        self.name = "memory"
        self.frequency = '1s'
        self.m_status_rss_max = {} 
        self.m_status_vms_max = {} 
        self.m_status_ipc_max = {} 
        self.m_status_ldsp_max = {} 
        self.m_status_llcp_max = {} 
        self.m_status_ipc_dec = {} 
        self.m_status_ldsp_inc = {} 
        self.m_status_llcp_inc = {} 
        self.m_status_gtbw_inc = {} 
        self.m_status_grbw_inc = {} 

        self.last_index = {}
        self.avg_window = 30 
        self.max_window = 30
        self.max_history = 60
        self.my_rank = config.wrank
        self.last_rcd = {}
        self.rss = {}
        self.vms = {} 
        self.ldsp = {}
        self.llcp = {}
        self.ipc = {}
        self.gtbw = {}
        self.grbw = {}
        self.fltrd_rss = {}
        self.fltrd_vms = {}
        self.fltrd_ldsp = {}
        self.fltrd_llcp = {}
        self.fltrd_ipc = {}
        self.fltrd_gtbw ={}
        self.fltrd_grbw ={}  
        self.urgent_update = False
        self.flags = {}          
        #if config.hc_lib  == 'papi':
        self.hd_counters = papi_counters[config.cpu_model.strip()]
        if 'gtbw' in self.hd_counters:
            self.hd_counters['gtbw'][0] = self.hd_counters['gtbw'][0].replace("DVN", 0)
        if 'grbw' in self.hd_counters:
            self.hd_counters['grbw'][0] = self.hd_counters['grbw'][0].replace("DVN", 0)

        self.metric_func = metrics[config.cpu_model.strip()]
        self.rss_thresh = math.ceil(0.90 * int(memory_size[config.cpu_model.strip()]))
        #print("[Rank ", self.my_rank, "] :","Machine name", config.cpu_model, "counter" , self.hd_counters)            
        self.update_model_conf(config)    

    def __compute_llc_miss_per(self, nprocs, llc_miss, llc_refs):
        llc_per, maxp = _percentage_('compute_llc_miss_per', nprocs, llc_miss, llc_refs)
        return llc_per
 
    def __compute_ld_stalls_per(self, nprocs, ld_stalls, core_cyc): 
        lds_per, maxp = _percentage_('compute_ld_stalls_per', nprocs, ld_stalls, core_cyc)
        return lds_per
        
    def __compute_ipc(self, nprocs, ins_ret, core_cyc): 
        ipc, maxv = _divide_('compute_ipc', nprocs, ins_ret, core_cyc)
        return ipc
    
    def __timestamp_to_date(self, unix_ts):
        ts_in_ms = [ j/1000000 for j in unix_ts]
        dateconv = numpy.vectorize(dt.datetime.fromtimestamp)
        date1 = dateconv(ts_in_ms)
        return date1.tolist()

    def __group_by_frequency(self, narray, by_last=0): #Replace S with U for microsecond 
        #print("[Rank ", self.my_rank, "] :",narray)  
        narray_v = narray[:,0] # get the value
        nlist_k = narray[:,1].tolist() #get the timestamp
        nlist_k = self.__timestamp_to_date(nlist_k)

        pdf = pd.DataFrame(data=narray_v, index=nlist_k, columns=["value"])
        #print("[Rank ", self.my_rank, "] :",pdf)
        if by_last == 0:
            pdf = pdf.groupby(pd.Grouper(freq=self.frequency)).sum()
        else:
            pdf = pdf.groupby(pd.Grouper(freq=self.frequency)).last()
        #pdf = pdf["value"].fillna(method='ffill')
        pdf = pdf["value"].dropna() #method='ffill')
        #print("[Rank ", self.my_rank, "] :",pdf)  
        return pdf #.to_dict()

    def get_model_name(self):
        return self.name

    def if_urgent_update(self):
        return self.urgent_update

    def update_model_conf(self, config): #active_reader_objs):
        self.adios2_active_conns = config.active_reader_objs
        self.r_map = config.local_res_map
        self.procs_per_ctr = config.reader_procs #range(30) #[ 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29] #config.reader_procs
        self.blocks_to_read = config.reader_blocks #[0] # {0: {"tau_metrics-lmp-mpi.bp": [0] } } #config.reader_blocks
        ini_val = 0 #need a postive value for the counter
        #print(self.blocks_to_read)
        ini_time = dt.datetime.now()
        pdf = pd.DataFrame(data=[ini_val], index=[ini_time], columns=["value"])
        pdf = pdf.groupby(pd.Grouper(freq=self.frequency)).last()  
        obj = numpy.array([[ini_val]])
        nodes = self.adios2_active_conns.keys()
        self.stream_prog = config.stream_prog
        for node in nodes:
            self.rss[node] = None
            self.vms[node] = None
            self.ldsp[node] = {}
            self.llcp[node] = {}
            self.ipc[node] = {}
            self.gtbw[node] = {}
            self.grbw[node] = {}
            self.fltrd_ldsp[node] = {}
            self.fltrd_ldsp[node] = {}
            self.fltrd_ldsp[node] = {}
            self.fltrd_llcp[node] = {}
            self.fltrd_ipc[node] = {}
            self.fltrd_ipc[node] = {}
            self.fltrd_gtbw[node] = {}
            self.fltrd_grbw[node] = {}
            self.fltrd_rss[node] = None 
            self.fltrd_vms[node] = None
            self.last_rcd[node] = {}
            self.m_status_rss_max[node] = []
            self.m_status_vms_max[node] = []
        
            self.m_status_ldsp_max[node] = {}
            self.m_status_llcp_max[node] = {}
            self.m_status_ipc_max[node] = {}
            self.m_status_ldsp_inc[node] = {}
            self.m_status_llcp_inc[node] = {}
            self.m_status_ipc_dec[node] = {}
            self.m_status_gtbw_inc[node] = {}
            self.m_status_grbw_inc[node] = {}
            self.last_index[node] = { 'rss':0, 'vms':0}
        
      
            streams = list(self.adios2_active_conns[node].keys())
            for stream in streams:
                procs = list(self.blocks_to_read[node][stream])
                #obj = OrderedCounter(pdf["value"].to_dict()) #numpy.array([[ini_time, ini_val]])
                #obj[ini_time] = ini_val
                self.last_index[node][stream] = {'lds':0, 'llc':0, 'ipc':0, 'gtbw':0, 'grbw':0}
                self.m_status_ldsp_max[node][stream] = []
                self.m_status_llcp_max[node][stream] = []
                self.m_status_ipc_max[node][stream] = []
                self.m_status_ldsp_inc[node][stream] = []
                self.m_status_llcp_inc[node][stream] = []
                self.m_status_ipc_dec[node][stream] = []
                self.m_status_grbw_inc[node][stream] = []
                self.m_status_grbw_inc[node][stream] = []

                self.last_rcd[node][stream] = {}
                for i in range(0,9): #for each metrics
                    self.last_rcd[node][stream][i] = obj 
                self.flags[MetricID.RSS] = True
                self.flags[MetricID.VMS] = True
                self.flags[MetricID.LLCM] = False
                self.flags[MetricID.LLCR] = False
                self.flags[MetricID.LDS] = False
                self.flags[MetricID.INS] = False
                self.flags[MetricID.CYC] = False
                self.flags[MetricID.GTBW] = False
                self.flags[MetricID.GRBW] = False
                for i in self.metric_func:
                    if i == 'llc_miss_per':
                        self.flags[MetricID.LLCM] = True
                        self.flags[MetricID.LLCR] = True
                    elif i == 'ld_stalls_per':     
                        self.flags[MetricID.LDS] = True
                        self.flags[MetricID.CYC] = True
                    elif i == 'ipc':     
                        self.flags[MetricID.INS] = True
                        self.flags[MetricID.CYC] = True
                    elif i == 'gtbw':     
                        self.flags[MetricID.GTBW] = True

                    elif i == 'grbw':     
                        self.flags[MetricID.GRBW] = True

                self.ldsp[node][stream] = None 
                self.fltrd_ldsp[node][stream] = None 
                self.llcp[node][stream] = None 
                self.fltrd_ldsp[node][stream] = None 
                self.ipc[node][stream] = None 
                self.gtbw[node][stream] = None 
                self.grbw[node][stream] = None 
                self.fltrd_ipc[node][stream] = None 
                self.fltrd_gtbw[node][stream] = None 
                self.fltrd_grbw[node][stream] = None 
    
    def __remove_empty_nan(self, array):
        mask = numpy.any(numpy.isnan(array), axis=1) # | numpy.equal(array, 0), axis=1)
        return array[~mask]

    def __get_agg_values_for(self, adios_conc, cntr, cntr_map, last_rcd, by_last=0, procs=[0], threads=[0], diff_idx=-1):
        is_valid, val_ar = adios_conc.read_var(cntr[0], procs, threads)
        if is_valid == True:
            for proc in procs:
                for th in threads:
                    val_tmp = val_ar[proc]
                    #tmp_ar = [val_tmp[:,0] == th] 
                    #print("[Rank ", self.my_rank, "] : process ", proc, " counter ", cntr, "cntr_map" , cntr, "Raw values ", val_tmp, flush = True)
                    #print("[Rank ", self.my_rank, "] :",val_tmp)
                    tmp_ar = val_tmp 
                    if tmp_ar.size != 0:
                        tmp_ar = tmp_ar[:,[2]]
                        #tmp_ar = self.__remove_empty_nan(tmp_ar)         
                        #print("[Rank ", self.my_rank, "] : process ", proc, " counter ", cntr, "After empty/Nan remova ", tmp_ar, flush = True)
  
                    if tmp_ar.size != 0:
                        #cntr_df = self.__group_by_frequency(tmp_ar, by_last)
                        #print("[Rank ", self.my_rank, "] : process ", proc, " counter ", cntr, "After group by ", cntr_df, flush = True)
                        cntr_df = tmp_ar #cntr_df.dropna()
                        if diff_idx != -1:
                             temp_df = last_rcd[diff_idx]
                             if cntr[1] == "C": 
                                 cntr_df = numpy.append(temp_df,tmp_ar)
                             #else:
                             #    cntr_df = temp_df
                             last_rcd[diff_idx] = cntr_df[-1:]
                             if cntr[1] == "C": 
                                cntr_df = numpy.diff(cntr_df)[0:]
                             #print("[Rank ", self.my_rank, "] : process ", proc, " counter ", cntr, "After diff ", cntr_df, flush = True)
                        if cntr_map is None:
                            cntr_map = cntr_df
                            #print("[Rank ", self.my_rank, "] :","Was None ", cntr_map )
                            #print("[Rank ", self.my_rank, "] : process " , proc , "counter ", cntr, " Counter mapping ", cntr_map, flush = True)
                        else:
                            cntr_map = numpy.concatenate((cntr_map, cntr_df), axis=0)
                            
                #print("[Rank ", self.my_rank, "] : process " , proc , "counter ", cntr, " Processed values ", cntr_map, flush = True)
        return cntr_map, last_rcd

    def update_curr_state(self):
        rss_val = {}
        vms_val = {}
        llcr_val = {} 
        llcm_val = {} 
        lds_val = {} 
        ins_val = {} 
        cyc_val = {}
        gtbw = {}
        grbw = {}
        k = 0 
        c_pressure = False
        b_pressure = False
        flag_llc = flag_lds = flag_ipc = False 
        nodes = self.adios2_active_conns.keys()
        #print("[Rank ", self.my_rank, "] :","Update for step ::", self.iter)
        self.iter = self.iter + 1  
        for node in nodes:
            rss_val = None
            vms_val = None
            #print("[Rank ", self.my_rank, "] :","Looking for node:: ", node)
            streams = list(self.adios2_active_conns[node].keys())
            for stream in streams:
                procs = list(self.blocks_to_read[node][stream])
                #print("[Rank ", self.my_rank, "] :","Looking for stream:: ", stream, " with procs", procs)
                #print("Active connections ", self.adios2_active_conns[node][stream]) 
                llcm_val = None
                llcr_val = None
                lds_val = None
                ins_val = None
                cyc_val = None
                gtbw = None
                grbw = None
                thread_l1 = [0]
                thread_l = [0,1,2,3]
                for active_conc in self.adios2_active_conns[node][stream]:
                #read RSS
                    rss_val, self.last_rcd[node][stream] = self.__get_agg_values_for(active_conc, self.rss_m, rss_val, self.last_rcd[node][stream], 1, procs, thread_l1)
                    #print("[Rank ", self.my_rank, "] :","Read RSS", self.last_rcd[node][stream][0])
                    #print("[Rank ", self.my_rank, "] :","Read RSS", rss_val)
                    vms_val, self.last_rcd[node][stream] = self.__get_agg_values_for(active_conc, self.vms_m, vms_val, self.last_rcd[node][stream], 1, procs, thread_l1)
                    #print("[Rank ", self.my_rank, "] :","Read VMS", self.last_rcd[node][stream][1])
                    #print("[Rank ", self.my_rank, "] :","Read VMS", vms_val)
                    #print("[Rank ", self.my_rank, "] :","Metric funcs are ", self.metric_func)
                    #print(rss_val) 
                    read_cyc = 0
                    for met in self.metric_func:
                        if met == "ld_stalls_per" or met == "ipc":
                            #read No of CPU cycles
                            if read_cyc != 1:
                                cyc_val, self.last_rcd[node][stream] = self.__get_agg_values_for(active_conc, self.hd_counters['cpu_cyc'], cyc_val, self.last_rcd[node][stream], 0, procs, thread_l1, MetricID.CYC.value)
                                read_cyc = 1
                                #print("[Rank ", self.my_rank, "] :","Read CYC:", cyc_val)
                        if met == "ld_stalls_per":
                            #read No of Load Stalls
                            lds_val, self.last_rcd[node][stream] = self.__get_agg_values_for(active_conc, self.hd_counters['ld_stalls'], lds_val, self.last_rcd[node][stream], 0, procs, thread_l1, MetricID.LDS.value)
                            #print("[Rank ", self.my_rank, "] :","Read STALLS: ", lds_val)
                        elif met == "llc_miss_per":
                            #read No of L3 misses
                            #print("[Rank ", self.my_rank, "] :",self.hd_counters['llc_misses'])
                            llcm_val, self.last_rcd[node][stream] = self.__get_agg_values_for(active_conc, self.hd_counters['llc_misses'], llcm_val, self.last_rcd[node][stream], 0, procs, thread_l1, MetricID.LLCM.value)
                            #print("[Rank ", self.my_rank, "] :","Read LLC MISS")
                            #print("[Rank ", self.my_rank, "] :READ LLCM",llcm_val)
                            #read No of L3 references
                            llcr_val, self.last_rcd[node][stream] = self.__get_agg_values_for(active_conc, self.hd_counters['llc_refs'], llcr_val, self.last_rcd[node][stream], 0, procs, thread_l1, MetricID.LLCR.value)
                            #print("[Rank ", self.my_rank, "] : READ LLCR",llcr_val)
                        elif met == "ipc":
                            #read No of Instruction
                            ins_val, self.last_rcd[node][stream] = self.__get_agg_values_for(active_conc, self.hd_counters['inst_ret'], ins_val, self.last_rcd[node][stream], 0, procs, thread_l1, MetricID.INS.value)
                            #print("[Rank ", self.my_rank, "] :","Read INS", ins_val)
                        elif met == "gtbw":
                            gtbw, self.last_rcd[node][stream] = self.__get_agg_values_for(active_conc, self.hd_counters['gtbw'], gtbw, self.last_rcd[node][stream], 0, procs, thread_l1, MetricID.GTBW.value)

                        elif met == "grbw":
                            grbw, self.last_rcd[node][stream] = self.__get_agg_values_for(active_conc, self.hd_counters['grbw'], grbw, self.last_rcd[node][stream], 0, procs, thread_l1, MetricID.GRBW.value)

                for met in self.metric_func:
                    if met == "llc_miss_per" and llcm_val is not None: 
                        if llcr_val is not None:
                            temp_rs = numpy.divide(numpy.sum(llcm_val, axis=0), numpy.sum(llcr_val, axis=0)) * 100
                        #temp_rs1 = temp_rs / temp_rs.shape[0]
                        #temp_rs1 = temp_rs1[temp_rs1[:,0] > 100]
                        #print(temp_rs, temp_rs.shape[0])
                            temp_rs = self.__filter_results(temp_rs, self.llcp[node], self.fltrd_llcp[node], stream) 
                            print("[Rank ", self.my_rank, "] :", "LLC VAR ", llcm_val, " LLCR  val ", llcr_val, " llc_per ", temp_rs,  flush=True)
                        #self.m_status_llcp_change[node][stream].append(self.__comput_ceange(self.m_status_llcp_max[node][stream]))
                        '''
                        self.m_status_llcp_max[node][stream].append(self.__compute_max(self.fltrd_llcp[node][stream]))
                        self.m_status_llcp_change[node][stream].append(self.__compute_inc(self.m_status_llcp_max[node][stream]))
                        if self.m_status_llcp_inc[node][stream][-1] >= 20:
                            flag_llc = True
                        '''
                    if met == "ld_stalls_per" and lds_val is not None:
                        if cyc_val is not None:
                            temp_rs = numpy.divide(numpy.sum(lds_val, axis = 0), numpy.sum(cyc_val, axis=0)) * 100
                        #print("LDS: ", lds_val )
                        #print("CYC: ", cyc_val )
                        #print("LDSP: ", temp_rs) 
                        #print(temp_rs.shape[0])
                        #temp_rs1 = temp_rs / temp_rs.shape[0]
                        #temp_rs1 = temp_rs1[temp_rs1[:,0] > 100]
                            temp_rs = self.__filter_results(temp_rs, self.ldsp[node], self.fltrd_ldsp[node], stream) 
                            print("[Rank ", self.my_rank, "] :","LDS VAR ", lds_val, " CYC  val ", cyc_val, " LDSP ", temp_rs,  flush=True)
                        '''
                        self.m_status_lds_inc[node][stream].append(self.__compute_inc(self.m_status_lds_max[node][stream]))
                        if self.m_status_llcp_inc[node][stream][-1] >= 20:
                            flag_lds = True
                        ''' 
                    if met == "ipc" and ins_val is not None:
                        if cyc_val is not None:
                            temp_rs = numpy.divide(numpy.sum(ins_val, axis = 0), numpy.sum(cyc_val, axis = 0))
                        #temp_rs = temp_rs / temp_rs.shape[0]
                            temp_rs = self.__filter_results(temp_rs, self.ipc[node], self.fltrd_ipc[node], stream) 
                            print("[Rank ", self.my_rank, "] :","INS VAR ", ins_val, " CYC  val ", cyc_val, " IPC ", temp_rs,  flush=True)
                        '''
                        self.m_status_ipc_max[node][stream].append(self.__compute_min(self.fltrd_ipc[node][stream]))
                        self.m_status_ipc_dec[node][stream].append(self.__compute_dec(self.m_status_ipc_max[node][stream]))
                         VAR ", llcm_val, " LLCR  val ", llcr_val, " llc_per ",
                            flag_ipc = True
                        '''
                    if met == "gtbw" and gtbw is not None:
                        if self.gtbw[node][stream] is None:
                            self.gtbw[node][stream] = gtbw 
                        else:
                            self.gtbw[node][stream] = numpy.concatenate((self.gtbw[node][stream], gtbw))

                    if met == "grbw" and grbw is not None:
                        if self.grbw[node][stream] is None:
                            self.grbw[node][stream] = grbw 
                        else:
                            self.grbw[node][stream] = numpy.concatenate((self.grbw[node][stream], grbw))

                    if flag_llc and flag_lds and flag_ipc:
                        b_pressure == True
            # keep only last X records
            if rss_val is not None:
               rss_val = (numpy.sum(rss_val, axis=0) / 1024) /1024
               rss_val = self.__filter_results(rss_val, self.rss, self.fltrd_rss, node, 1 ) 
               print("[Rank ", self.my_rank, "] :", "RSS VAR ", rss_val, flush=True)

            if vms_val is not None:
               vms_val = (numpy.sum(vms_val, axis=0) / 1024) /1024
               vms_val = self.__filter_results(vms_val, self.vms, self.fltrd_vms, node, 1 ) 

        self.dump_curr_state()
        return True

    def __filter_results(self, arr, res_dict, res_fltrd, idx, axis = 0):
        if arr is not None:
           if axis != 1:
               arr = arr[ ~numpy.isnan(arr)]
           else:
               arr = [ numpy.sum(arr, axis=0)]
            #arr1 = arr[numpy.logical_not(numpy.isnan(arr))]

        if arr is not None:
            if idx not in res_dict.keys() or res_dict[idx] is None: 
                res_dict[idx] = arr
            else:
                print(res_dict[idx], arr) 
                res_dict[idx] = numpy.concatenate([res_dict[idx], arr])
        else:
            res_dict[idx] = None

        #print("[Rank ", self.my_rank, "] :","Filtered RESULT::\n", res_dict[idx])
        if res_dict is not None and res_dict[idx] is not None:# update the new state
            #res_fltrd[idx] = self.__compute_change(res_dict[idx][-self.max_history:])
            res_fltrd[idx] = self.__compute_ravg(res_dict[idx][-self.max_history:])

            #if res_fltrd is not None:
            #    self.m_status_vms_max[node] = numpy.concatenate((self.m_status_vms_max[node], [self.__compute_max(res_fltrd)] ))
        #print("[Rank ", self.my_rank, "] :","Filtered RESULT::\n", res_fltrd[idx])
        return arr    

    def __compute_ravg(self, arr):
        if arr is None:
            return None 
        df = pd.DataFrame(arr)
        df_avg = df.rolling(self.avg_window, min_periods=1).mean()       
        return df_avg.to_numpy()
    
    def __compute_rmax(self, arr):
        if arr is None:
            return None 
        df = pd.DataFrame(arr)
        df_max = df.rolling(self.max_window, min_periods=1).max()       
        return df_max.to_numpy()

    def __compute_max(self, array):
        if isinstance(array,dict):
            return 0
        if array is None:
            return 0 
        array_max = array.max()
        return array_max

    def __compute_min(self, array):
        if isinstance(array,dict):
            return 0
        if array is None:
            return 0 
        #print("[Rank ", self.my_rank, "] :",df.head())
        array_min = array.min()
        return array_min

    def __compute_inc(self, lst):
        if lst is None:
            return 0 
        val_n = lst[-1]
        val_o = lst[-2]
        return ((val_n - val_o)/val_o) * 100 

    def __compute_dec(self, lst):
        if lst is None:
            return 0 
        val_n = lst[-1]
        val_o = lst[-2]
        return ((val_o - val_n)/val_o) * 100 

    def __compute_change(self, lst):
        if lst is None:
            return 0
        val = lst[0]
        sum_c = 0
        for i in range(1, len(lst)): 
            sum_c += (lst[i] - val) #**2
        #return ((val - sum_c/len(lst))/val) * 100   
        return  sum_c/len(lst)   

    def __check_overflow(self, ar, writer, index, sheetnm):
        if ar is None:
            return
        total_rows = numpy.shape(ar)[0] #ar.shape[0] 
        #print("[Rank ", self.my_rank, "] :",total_rows)
        if total_rows > index:
            temp_ar = ar[index:]
            #if temp_df.shape[0] > 1:
                #temp_df = temp_df[-2:]
            pd.DataFrame(temp_ar).to_csv(writer, header = False)
            #print("[Rank ", self.my_rank, "] : WROTE ", temp_ar, flush=True)
            #temp_df.to_excel(xlsx_writer, startrow = index, sheet_name = sheetnm)
            if self.max_history > total_rows:
                ar = ar[-self.max_history:]
 
            total_rows = numpy.shape(ar)[0] 
            index = total_rows
        return ar, index
 
    def create_workbook(self, filename, sheetnames):
        workbook = xlsxwriter.Workbook(filename)
        for sheet in sheetnames:
            worksheet = workbook.add_worksheet(sheet)
        workbook.close()

    def dump_curr_state(self):
        nodes = self.adios2_active_conns.keys()
        file_md = 'a'
        #print("[Rank ", self.my_rank, "] :","NODES..", nodes)
        for node in nodes:
            filename = "memory-" + str(node) + "-rss.csv"
            print("[Rank ", self.my_rank, "] :","WRITING..", filename) 
            print("[Rank ", self.my_rank, "] :","WRITING..RSS to", filename, self.rss[node]) 
            if self.rss[node] is not None:
                with open(filename, 'a') as writer:
                     self.rss[node], self.last_index[node]['rss'] = self.__check_overflow(self.rss[node], writer, self.last_index[node]['rss'], 'rss')
            filename = "memory-" + str(node) + "-vms.csv"
            #print("[Rank ", self.my_rank, "] :","WRITING..", filename) 
            if self.vms[node] is not None:
                with open(filename, 'a') as writer:
                    self.vms[node], self.last_index[node]['vms'] = self.__check_overflow(self.vms[node], writer, self.last_index[node]['vms'], 'vms')

            streams = list(self.adios2_active_conns[node].keys())
            #print(self.flags)
            for stream in streams:
                if self.flags[MetricID.LLCM] and self.llcp[node][stream] is not None:
                    filename = "memory-" + str(node) + "-" + str(stream.replace('.', '').replace('/', '-')) + "-llcp.csv"
                    with open(filename, 'a') as writer:
                        self.llcp[node][stream], self.last_index[node][stream]['llc'] = self.__check_overflow(self.llcp[node][stream], writer, self.last_index[node][stream]['llc'], 'llc')
                if self.flags[MetricID.LDS] and self.ldsp[node][stream] is not None:
                    filename = "memory-" + str(node) + "-" + str(stream.replace('.', '').replace('/', '-')) + "-ldsp.csv"
                    with open(filename, 'a') as writer:
                        self.ldsp[node][stream], self.last_index[node][stream]['lds'] = self.__check_overflow(self.ldsp[node][stream], writer, self.last_index[node][stream]['lds'], 'lds')
                if self.flags[MetricID.INS] and self.ipc[node][stream] is not None:
                    filename = "memory-" + str(node) + "-" + str(stream.replace('.', '').replace('/', '-')) + "-ipc.csv"
                    with open(filename, 'a') as writer:
                        self.ipc[node][stream], self.last_index[node][stream]['ipc'] = self.__check_overflow(self.ipc[node][stream], writer, self.last_index[node][stream]['ipc'], 'ipc')
                if self.flags[MetricID.GTBW] and self.gtbw[node][stream] is not None:
                    filename = "memory-" + str(node) + "-" + str(stream.replace('.', '').replace('/', '-')) + "-gtbw.csv"
                    with open(filename, 'a') as writer:
                        self.gtbw[node][stream], self.last_index[node][stream]['gtbw'] = self.__check_overflow(self.gtbw[node][stream], writer, self.last_index[node][stream]['gtbw'], 'gtbw')
                if self.flags[MetricID.GRBW] and self.grbw[node][stream] is not None:
                    filename = "memory-" + str(node) + "-" + str(stream.replace('.', '').replace('/', '-')) + "-grbw.csv"
                    with open(filename, 'a') as writer:
                        self.grbw[node][stream], self.last_index[node][stream]['grbw'] = self.__check_overflow(self.grbw[node][stream], writer, self.last_index[node][stream]['grbw'], 'grbw')
        return   
    

    def get_curr_state(self):
        j_data = {}
        j_data['RSS'] = {}
        j_data['VMS'] = {}
        j_data['RSS']['NODE-WORKFLOW'] = [] #self.rss[node][-1][0] #self.m_status_rss_max[node][-1] 
        j_data['VMS']['NODE-WORKFLOW'] = [] #self.rss[node][-1][0] #self.m_status_rss_max[node][-1] 
        j_data['LLC'] = {}
        j_data['LDS'] = {}
        j_data['IPC'] = {}
        j_data['LLC']['NODE-TASK'] = {}
        j_data['LDS']['NODE-TASK'] = {}
        j_data['IPC']['NODE-TASK'] = {}
        nodes = self.adios2_active_conns.keys()
        for node in nodes:
            j_data['RSS']['NODE-WORKFLOW'].append(self.rss[node][-1]) #self.rss[node][-1][0] #self.m_status_rss_max[node][-1] 
            j_data['VMS']['NODE-WORKFLOW'].append(self.vms[node][-1])  #self.vms[node][-1][0] #self.m_status_vms_max[node][-1] 
            gran = 'NODE-TASK'
            streams = list(self.adios2_active_conns[node].keys())
            for stream in streams:
                #if len(self.fltrd_llcp[node][stream]) == 0: # self.m_status_llcp_inc[node][stream]) == 0:
                #    j_data[node]['LLC'][stream] = []
                #else:  
                    #j_data[node]['LLC'][stream] = self.m_status_llcp_inc[node][stream][-1]
                strprog = self.stream_prog[node][stream] 
                if stream not in j_data['LLC'][gran].keys():
                    j_data['LLC'][gran][strprog] = [] 
                j_data['LLC'][gran][strprog].append(self.fltrd_llcp[node][stream][-1][0]) #m_status_llcp_inc[node][stream][-1] 
                #if len(self.fltrd_ldsp[node][stream]) == 0: # self.m_status_ldsp_inc[node][stream]) == 0:
                #    j_data[node]['LDS'][stream] = []
                #else:  
                if stream not in j_data['LDS'][gran].keys():
                    j_data['LDS'][gran][strprog] = [] 
                j_data['LDS'][gran][strprog].append(self.fltrd_ldsp[node][stream][-1][0]) #[-1] #self.m_status_ldsp_inc[node][stream][-1] 
                #if len(self.fltrd_ipc[node][stream]) == 0 : #self.m_status_ipc_dec[node][stream]) == 0:
                #    j_data[node]['IPC'][stream] = []
                #else:  
                if stream not in j_data['IPC'][gran].keys():
                    j_data['IPC'][gran][strprog] = [] 
                j_data['IPC'][gran][strprog].append(self.fltrd_ipc[node][stream][-1][0]) #[-1] #self.m_status_ipc_dec[node][stream][-1] 
        #self.dump_curr_state()
        return { self.my_rank : j_data  } 

    def merge_curr_state(self, js_data):
        m_data = {}
        m_data['RSS'] = {}
        m_data['VMS'] = {}
        m_data['RSS']['NODE-WORKFLOW'] = [] #self.rss[node][-1][0] #self.m_status_rss_max[node][-1] 
        m_data['VMS']['NODE-WORKFLOW'] = [] #self.rss[node][-1][0] #self.m_status_rss_max[node][-1] 
        m_data['LLC'] = {}
        m_data['LDS'] = {}
        m_data['IPC'] = {}
        m_data['LLC']['NODE-TASK'] = {}
        m_data['LDS']['NODE-TASK'] = {}
        m_data['IPC']['NODE-TASK'] = {}
        js_data = js_data[0]        
        for proc in js_data.keys():
            j_data = js_data[proc]
            m_data['RSS']['NODE-WORKFLOW'].append(j_data['RSS']['NODE-WORKFLOW']) #self.rss[node][-1][0] #self.m_status_rss_max[node][-1] 
            m_data['VMS']['NODE-WORKFLOW'].append(j_data['RSS']['NODE-WORKFLOW'])
            gran = 'NODE-TASK'
            streams = j_data['LLC'][gran].keys()
            for stream in streams:
                if stream not in m_data['LLC'][gran].keys():
                    m_data['LLC'][gran][stream] = []  
                m_data['LLC'][gran][stream].append(j_data['LLC'][gran][stream]) #m_status_llcp_inc[node][stream][-1] 
                #if len(self.fltrd_ldsp[node][stream]) == 0: # self.m_status_ldsp_inc[node][stream]) == 0:
                #    j_data[node]['LDS'][stream] = []
                #else:  
                if stream not in m_data['LDS'][gran].keys():
                    m_data['LDS'][gran][stream] = []  
                m_data['LDS'][gran][stream].append(j_data['LDS'][gran][stream]) #[-1] #self.m_status_ldsp_inc[node][stream][-1] 
                #if len(self.fltrd_ipc[node][stream]) == 0 : #self.m_status_ipc_dec[node][stream]) == 0:
                #    j_data[node]['IPC'][stream] = []
                #else:  
                if stream not in m_data['IPC'][gran].keys():
                    m_data['IPC'][gran][stream] = []  
                m_data['IPC'][gran][stream].append(j_data['IPC'][gran][stream]) #[-1] #self.m_status_ipc_dec[node][stream][-1] 
        #self.dump_curr_state()   
        return m_data        


class config():
   def __init__(self, model, adios_reader, adios_reader_blk):
       self.cpu_model = model
       self.active_reader_objs = adios_reader
       self.local_res_map = {}
       self.reader_procs = range(30) #[ 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29] #config.reader_procs
       self.reader_blocks = adios_reader_blk
       self.wrank = 0

if __name__ == "__main__":
    adios_reader = {}
    adios_reader_blks  = {}
    nstreams = int(sys.argv[1])
    node = "1"
    adios_reader[node] = {}
    adios_reader_blks[node]  = {}
    nprocs = {}
    streams = {}
    
    for i in range(nstreams): 
        streams[i] = str(sys.argv[i+2])
        adios_reader[node][streams[i]] = []
        adios_reader_blks[node][streams[i]] = [0]

    for i in range(nstreams): 
        nprocs[i] = range(int(sys.argv[i + 2 + nstreams]))

    #gpu_n = sys.argv[nstreams + 2 + nstreams]
 
    '''
    stream1 = "simulation/tauprofile-lmp_mpi.bp" #str(sys.argv[1]) # tau-metric stream name (application) to connect..
    stream2 = "cna_calc/tauprofile-lmp_mpi.bp" #str(sys.argv[1]) # tau-metric stream name (application) to connect..
    procs1 = range(30) #list(sys.argv[2]) # processes (of this application) that wrote seperate streams on this node..
    procs2 = range(10) #list(sys.argv[2]) # processes (of this application) that wrote seperate streams on this node..
    adios_reader[node][stream1] = []
    adios_reader[node][stream2] = []
    '''
    adios2_engine = "BP4" #str(sys.argv[3]) # adios2 engine to use for reading..
    freq_sec = "5s" #sys.argv[4] # frequency at which the data needs to be organised (in sec) - minimum value 1s..
  
    for i in range(nstreams):
        for proc in nprocs[i]:
            str_split = streams[i].split('.bp')
            con_str = str_split[0] + "-" + str(proc) + ".bp"
            reader_obj = adios2_tau_reader(con_str, adios2_engine, MPI.COMM_WORLD, [0], 'trace')
            adios_reader[node][streams[i]].append(reader_obj)

    cpu_info = "Intel(R) Xeon(R) CPU E5-2680 v2 @ 2.80GHz" #Intel(R) Xeon(R) Gold 6248 CPU @ 2.50GHz"

    cnf = config(cpu_info, adios_reader, adios_reader_blks) 

    m_stats = memory(cnf) #, gpu_n)

    reader_objs = [1]
    
    for i in range(nstreams):
        reader_objs.extend(adios_reader[node][streams[i]])

    reader_objs = reader_objs[1:]
    done = 0

    for reader_ob in reader_objs:
        reader_ob.open()

    while done == 0:
        done = 1
        for reader_ob in reader_objs:
            #print (reader_ob)
            ret = reader_ob.advance_step()
            if ret == True:
               done = 0

        if done == 0:
            m_stats.update_curr_state()

            for reader_ob in reader_objs:
                reader_ob.end_step()
        #done = 1
    for reader_ob in reader_objs:
        reader_ob.close()
    j_str = m_stats.get_curr_state()
    print(" ****** ")
    print(j_str)
    print(" ****** ")
    m_stats.dump_curr_state()
